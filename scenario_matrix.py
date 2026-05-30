#!/usr/bin/env python3
"""Run many retirement scenarios from one Excel sheet.

Column A of the sheet lists input parameters (friendly labels, grouped by
section). Columns B, C, D... are scenarios. Column B is the Baseline; any blank
cell in a later column inherits the Baseline value, so each scenario only needs
the cells that change.

Two subcommands:

    # 1. Generate a template workbook seeded from a base plan
    python3 scenario_matrix.py template --base base.json --out scenario_matrix.xlsx [--scenarios 3]

    # 2. Fill in the workbook, then run every column and print a comparison
    python3 scenario_matrix.py run scenario_matrix.xlsx --base base.json [--runs 1000]

The base plan is a Streamlit-format plan JSON (same shape as temp_roth_80eq.json).
Privacy: the base-plan path is passed on the command line and never stored in the
workbook; this tool never prints the plan's `client` field.
"""
import argparse
import copy
import json
import sys
from pathlib import Path

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).resolve().parent))

from compare_modes import _metrics, _run_one  # reuse the headless engine glue
from sim_engine import BOOTSTRAP_MODE_NAME

# Friendly return-mode words -> exact engine strings used by _run_one().
RETURN_MODE_MAP = {
    "historical": "Historical (master_global_factors)",
    "lognormal": "Simulated (lognormal)",
    "simulated": "Simulated (lognormal)",
    "bootstrap": BOOTSTRAP_MODE_NAME,
}

SHEET_NAME = "Scenarios"
NOTE_TEXT = "Every column starts at the same defaults — change only the cells you want."


# --------------------------------------------------------------------------- #
# Field map: friendly label -> how to read/write it on a flat plan dict.       #
# --------------------------------------------------------------------------- #

def _set_spending(plan, value):
    """Spending lives at periods[0].amount, not a flat key."""
    periods = plan.get("periods")
    if not periods:
        periods = [{"amount": 0.0}]
        plan["periods"] = periods
    periods[0]["amount"] = float(value)


def _get_spending(plan):
    periods = plan.get("periods") or [{}]
    return periods[0].get("amount")


class Field:
    def __init__(self, label, key, kind, choices=None, get_fn=None, set_fn=None):
        self.label = label
        self.key = key
        self.kind = kind  # int | dollar | pct | fraction | bool | choice
        self.choices = choices
        self.get_fn = get_fn
        self.set_fn = set_fn

    def read(self, plan):
        if self.get_fn:
            return self.get_fn(plan)
        return plan.get(self.key)

    def coerce(self, raw):
        """Turn a spreadsheet cell value into the plan-ready value."""
        if self.kind in ("int",):
            return int(round(float(raw)))
        if self.kind in ("dollar", "num"):
            return float(raw)
        if self.kind == "pct":
            # Stored 0-100 in the plan JSON (translator divides by 100).
            return float(raw)
        if self.kind == "fraction":
            v = float(raw)
            return v / 100.0 if v > 1 else v
        if self.kind == "bool":
            s = str(raw).strip().lower()
            return s in ("1", "true", "yes", "y", "t")
        if self.kind == "choice":
            s = str(raw).strip()
            if self.key == "return_mode":
                m = RETURN_MODE_MAP.get(s.lower())
                if m is None:
                    raise ValueError(
                        f"Unknown return mode {s!r}; use one of "
                        f"{sorted(set(k.title() for k in RETURN_MODE_MAP))}"
                    )
                return m
            # Case-insensitive match against allowed friendly values.
            for c in self.choices or []:
                if s.lower() == c.lower():
                    return c
            raise ValueError(f"Unknown value {s!r} for {self.label!r}; allowed: {self.choices}")
        return raw

    def apply(self, plan, raw):
        value = self.coerce(raw)
        if self.set_fn:
            self.set_fn(plan, value)
        else:
            plan[self.key] = value


# Section header -> ordered list of Fields. Drives both template and run.
SECTIONS = [
    ("PEOPLE", [
        Field("Person 1 starting age", "start_age", "int"),
        Field("Person 1 life expectancy", "life_expectancy_primary", "int"),
        Field("Person 2 starting age", "start_age_spouse", "int"),
        Field("Person 2 life expectancy", "life_expectancy_spouse", "int"),
    ]),
    ("ACCOUNTS", [
        Field("Taxable balance", "taxable_start", "dollar"),
        Field("TDA — Person 1", "tda_start", "dollar"),
        Field("TDA — Person 2", "tda_spouse_start", "dollar"),
        Field("Roth balance", "roth_start", "dollar"),
    ]),
    ("SPENDING", [
        Field("Annual spending (after-tax)", "__spending__", "dollar",
              get_fn=_get_spending, set_fn=_set_spending),
    ]),
    ("ALLOCATION", [
        Field("Stock %", "target_stock_pct", "pct"),
        Field("Withdraw TDA before taxable", "prefer_tda_before_taxable", "bool"),
    ]),
    ("SOCIAL SECURITY", [
        Field("SS benefit — Person 1", "ss_income", "dollar"),
        Field("SS claim age — Person 1", "ss_start_age_p1", "int"),
        Field("SS benefit — Person 2", "ss_income_spouse", "dollar"),
        Field("SS claim age — Person 2", "ss_start_age_p2", "int"),
    ]),
    ("PENSIONS", [
        Field("Pension income — Person 1", "pension_income", "dollar"),
        Field("Pension income — Person 2", "pension_income_spouse", "dollar"),
    ]),
    ("ROTH CONVERSIONS", [
        Field("Roth conversion mode", "roth_conversion_mode", "choice",
              choices=["None", "Fixed amount", "Fill to bracket"]),
        Field("Roth conversion amount", "roth_conversion_amount", "dollar"),
        Field("Roth conversion years", "roth_conversion_years", "int"),
        Field("Roth bracket fill rate", "roth_bracket_fill_rate", "fraction"),
    ]),
    ("TAXES", [
        Field("Filing status", "filing_status", "choice",
              choices=["Single", "Married Filing Jointly"]),
        Field("State tax rate", "state_tax_rate", "fraction"),
        Field("State exempts retirement income", "state_exempt_retirement", "bool"),
    ]),
    ("GUARDRAILS", [
        Field("Guardrails enabled", "guardrails_enabled", "bool"),
        Field("Lower guardrail", "guardrail_lower", "fraction"),
        Field("Upper guardrail", "guardrail_upper", "fraction"),
        Field("Target success", "guardrail_target", "fraction"),
    ]),
    ("RETURNS", [
        Field("Return mode", "return_mode", "choice",
              choices=["Historical", "Lognormal", "Bootstrap"]),
        Field("Investment fee (basis points)", "investment_fee_bps", "num"),
    ]),
    ("SIMULATION", [
        Field("Monte Carlo runs", "monte_carlo_runs", "int"),
    ]),
]

# Flat label -> Field lookup (label is the spreadsheet key in column A).
FIELD_BY_LABEL = {f.label: f for _, fields in SECTIONS for f in fields}


def _load_plan(path):
    return json.loads(Path(path).read_text())


def _display_value(field, plan):
    """Human-friendly value for column B of the template."""
    val = field.read(plan)
    if val is None:
        return None
    if field.kind == "choice" and field.key == "return_mode":
        # Reverse-map the engine string to a friendly word, if possible.
        for friendly, engine in RETURN_MODE_MAP.items():
            if engine == val:
                return friendly.title()
        return val
    if field.kind == "bool":
        return bool(val)
    return val


# --------------------------------------------------------------------------- #
# template subcommand                                                          #
# --------------------------------------------------------------------------- #

def cmd_template(args):
    plan = _load_plan(args.base)
    n_scen = max(1, args.scenarios)

    # Starting defaults applied to EVERY column so each scenario begins from the
    # same known point — the user then changes only the cells they care about.
    seed = copy.deepcopy(plan)
    seed["target_stock_pct"] = float(args.equity)
    seed["investment_fee_bps"] = float(args.fee_bps)
    seed["return_mode"] = RETURN_MODE_MAP[args.return_mode.lower()]

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    bold = Font(bold=True)
    section_fill = PatternFill("solid", fgColor="DDE7F0")
    note_font = Font(italic=True, color="808080")

    # Row 1: note. Row 2: header. Data starts row 3.
    ws.cell(row=1, column=1, value=NOTE_TEXT).font = note_font

    headers = ["Input"] + [f"Scenario {i}" for i in range(1, n_scen + 1)]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = bold
        c.alignment = Alignment(horizontal="left")

    last_data_col = n_scen + 1  # columns 2..last_data_col hold scenario values
    row = 3
    dropdown_rows = []  # (row, [allowed values]) for choice/bool fields
    for section, fields in SECTIONS:
        sc = ws.cell(row=row, column=1, value=section)
        sc.font = bold
        sc.fill = section_fill
        row += 1
        for f in fields:
            ws.cell(row=row, column=1, value=f.label)
            bval = _display_value(f, seed)
            if bval is not None:
                for col in range(2, last_data_col + 1):
                    ws.cell(row=row, column=col, value=bval)
            if f.kind == "choice":
                dropdown_rows.append((row, list(f.choices)))
            elif f.kind == "bool":
                dropdown_rows.append((row, ["TRUE", "FALSE"]))
            row += 1

    # Dropdowns (Excel data validation) across all scenario columns B..last.
    last_col = get_column_letter(last_data_col)
    for drow, values in dropdown_rows:
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(values) + '"',
            allow_blank=True,
            showDropDown=False,  # False = show the arrow (Excel quirk)
        )
        ws.add_data_validation(dv)
        dv.add(f"B{drow}:{last_col}{drow}")

    # Formatting: widths + freeze header row and label column.
    ws.column_dimensions["A"].width = 34
    for i in range(2, last_data_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.freeze_panes = "B3"

    out = Path(args.out)
    wb.save(out)
    print(f"Wrote template: {out}")
    print(f"  {len(FIELD_BY_LABEL)} inputs x {n_scen} scenario columns, all seeded to defaults")
    print(f"  ({args.equity:.0f}% equity, {args.fee_bps:.0f} bps fee, {args.return_mode} returns).")
    print(f"  Edit only the cells you want, then run:")
    print(f"    python3 scenario_matrix.py run {out} --base {args.base}")


# --------------------------------------------------------------------------- #
# run subcommand                                                              #
# --------------------------------------------------------------------------- #

def _build_scenarios(ws, base_plan):
    """Return (names, plans, modes) from the worksheet."""
    header_row = None
    for r in range(1, 6):  # header is near the top; tolerate a note row above it
        if (ws.cell(row=r, column=1).value or "").strip().lower() == "input":
            header_row = r
            break
    if header_row is None:
        raise SystemExit("Could not find the header row (expected 'Input' in column A).")

    # Scenario columns: every column >= 2 with a non-empty header.
    scen_cols = []
    col = 2
    while True:
        h = ws.cell(row=header_row, column=col).value
        if h is None or str(h).strip() == "":
            # allow a single gap then stop
            if ws.cell(row=header_row, column=col + 1).value:
                col += 1
                continue
            break
        scen_cols.append((col, str(h).strip()))
        col += 1
    if not scen_cols:
        raise SystemExit("No scenario columns found (need a header in column B or later).")

    names = [name for _, name in scen_cols]
    plans = [copy.deepcopy(base_plan) for _ in scen_cols]
    modes = [None] * len(scen_cols)
    unknown = set()

    for r in range(header_row + 1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if label is None:
            continue
        label = str(label).strip()
        field = FIELD_BY_LABEL.get(label)
        if field is None:
            # Section headers and free-text rows are fine; only warn on rows
            # that actually carry scenario values.
            if any(ws.cell(row=r, column=c).value not in (None, "") for c, _ in scen_cols):
                unknown.add(label)
            continue
        for idx, (c, _name) in enumerate(scen_cols):
            raw = ws.cell(row=r, column=c).value
            if raw is None or (isinstance(raw, str) and raw.strip() == ""):
                continue  # blank -> inherit baseline
            try:
                field.apply(plans[idx], raw)
            except (ValueError, TypeError) as e:
                raise SystemExit(f"Scenario '{names[idx]}', row '{label}': {e}")
            if field.key == "return_mode":
                modes[idx] = plans[idx]["return_mode"]

    # Default each scenario's mode to its (possibly inherited) plan return_mode.
    for idx in range(len(modes)):
        if modes[idx] is None:
            modes[idx] = plans[idx].get("return_mode", "Simulated (lognormal)")

    if unknown:
        print("Warning: ignoring unrecognized input rows: " + ", ".join(sorted(unknown)))
    return names, plans, modes


def _fmt(v, kind):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "—"
    if kind == "int":
        return f"{int(v):,}"
    if kind == "pct":
        return f"{v:.1f}%"
    if kind == "dollar":
        return f"${v:,.0f}"
    return str(v)


# (metric_key, row label, format kind) — same metrics compare_modes surfaces.
METRIC_ROWS = [
    ("n_paths", "Paths", "int"),
    ("portfolio_survival_%", "Survival %", "pct"),
    ("spending_success_%", "Spending success %", "pct"),
    ("mean_CAGR_%", "Mean CAGR", "pct"),
    ("median_CAGR_%", "Median CAGR", "pct"),
    ("P5_after_tax_end", "P5 ending", "dollar"),
    ("P10_after_tax_end", "P10 ending", "dollar"),
    ("P50_after_tax_end", "P50 ending", "dollar"),
    ("P90_after_tax_end", "P90 ending", "dollar"),
    ("mean_spending", "Mean avg-annual spending", "dollar"),
    ("P10_spending", "P10 avg-annual spending", "dollar"),
    ("P50_spending", "P50 avg-annual spending", "dollar"),
]


def cmd_run(args):
    base_plan = _load_plan(args.base)
    wb = load_workbook(args.xlsx, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    names, plans, modes = _build_scenarios(ws, base_plan)

    print(f"Scenarios: {len(names)}   Monte Carlo paths: {args.runs}\n")
    metrics = []
    headers_info = []
    for name, plan, mode in zip(names, plans, modes):
        friendly_mode = next(
            (k.title() for k, v in RETURN_MODE_MAP.items() if v == mode), mode
        )
        print(f"Running {name} ({friendly_mode}) ...", flush=True)
        results, ayd, sim_years, base_spending = _run_one(plan, mode, args.runs)
        m = _metrics(results, ayd, sim_years, base_spending)
        metrics.append(m)
        stock = float(plan.get("target_stock_pct", 0))
        headers_info.append(
            f"{sim_years}y / {stock:.0f}% stk / ${base_spending:,.0f} / {friendly_mode}"
        )

    # Side-by-side table: metrics as rows, scenarios as columns.
    label_w = max(len(lbl) for _, lbl, _ in METRIC_ROWS) + 2
    col_w = max(18, max(len(n) for n in names) + 2)
    sep = "  "

    print()
    print(sep.join(["".ljust(label_w)] + [n.rjust(col_w) for n in names]))
    print(sep.join(["".ljust(label_w)] + [info.rjust(col_w) for info in
          [h[:col_w] for h in headers_info]]))
    print(sep.join(["-" * label_w] + ["-" * col_w] * len(names)))
    for key, lbl, kind in METRIC_ROWS:
        cells = []
        for m in metrics:
            cells.append(_fmt(m.get(key) if m else None, kind).rjust(col_w))
        print(sep.join([lbl.ljust(label_w)] + cells))
    return 0


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)

    t = sub.add_parser("template", help="Generate a scenario-matrix workbook from a base plan")
    t.add_argument("--base", required=True, help="Path to base plan JSON (Streamlit/session-state schema)")
    t.add_argument("--out", default="scenario_matrix.xlsx", help="Output .xlsx path")
    t.add_argument("--scenarios", type=int, default=5,
                   help="Number of scenario columns, all seeded to defaults (default 5)")
    t.add_argument("--equity", type=float, default=60.0,
                   help="Starting stock %% applied to every column (default 60)")
    t.add_argument("--fee-bps", type=float, default=100.0,
                   help="Starting investment fee in basis points (default 100)")
    t.add_argument("--return-mode", default="Historical",
                   choices=["Historical", "Lognormal", "Bootstrap"],
                   help="Starting return mode applied to every column (default Historical)")
    t.set_defaults(func=cmd_template)

    r = sub.add_parser("run", help="Run every scenario column and print a comparison")
    r.add_argument("xlsx", help="Path to the filled scenario-matrix workbook")
    r.add_argument("--base", required=True, help="Path to base plan JSON (blanks inherit from it)")
    r.add_argument("--runs", type=int, default=1000, help="Monte Carlo paths per stochastic scenario")
    r.set_defaults(func=cmd_run)

    args = ap.parse_args()
    return args.func(args) or 0


if __name__ == "__main__":
    raise SystemExit(main())
